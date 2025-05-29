import pandas as pd
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def data_center(new_file_path, insert_cell):
    wb = load_workbook(new_file_path)
    ws = wb.active
    align_center = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align_center
    wb.save(new_file_path)
    print(f"新文件 {new_file_path} 已创建，数据写入位置：{insert_cell}")


def show_info(group, new_file_path, insert_cell, start_date, team_name):
    work_info_list = []
    for data ,day_work in enumerate(group):
        for number, workshop in enumerate(day_work, 1):
            for work in workshop:

                work.real_start_sand_time = (timedelta(days=data) + start_date).date()
                work.real_end_sand_time = (timedelta(days=data + 1) + start_date).date()
                work.pre_end_paint_time = work.pre_end_paint_time.date()

                work_info = [work.GC ,work.ZD ,work.SG, work.sand_area, work.paint_area,
                             work.ZYQ, team_name[number - 1],
                             work.real_start_sand_time,
                             work.real_end_sand_time,
                             work.CX, work.length, work.weight, work.height, work.n_paint]
                work_info_list.append(work_info)

    work_info_list.insert(0 ,["工程号" ,"总段号" ,"施工项目" ,"冲砂面积" ,"涂漆面积",
                             "冲砂作业区" ,"跨间班组" ,"开始时间" ,"结束时间",
                             "船型" ,"分段长" ,"分段宽" ,"分段高" ,"涂漆次数"])
    # 将数据转为DataFrame
    df = pd.DataFrame(work_info_list)

    # 计算起始行列索引（从0开始）
    start_row = int(insert_cell[1:]) - 1
    start_col = ord(insert_cell[0].upper()) - ord("A")

    # 直接写入新文件
    with pd.ExcelWriter(new_file_path, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            index=False,
            header=False,
            startrow=start_row,
            startcol=start_col
        )
    data_center(new_file_path, insert_cell)


def write_group_assignment(total_paint_list, new_file_path, insert_cell, team_name, paint_efficiency):

    data_list = []

    data_list.append(["工程号" ,"总段号" ,"施工项目" ,"冲砂面积" ,"涂漆面积",
                      "涂漆作业区" ,"跨间班组" ,"喷涂班组", "冲砂结束",
                      "船型" ,"分段长" ,"分段宽" ,"分段高" ,"涂漆次数", "预计涂漆结束", "实际涂漆结束"])

    for workshop_idx, groups in enumerate(total_paint_list, 0):  # 车间从0开始编号
        for group_idx, segments in enumerate(groups, 1):  # 小组从1开始编号
            segments = sorted(segments, key=lambda x:x.real_end_sand_time)
            start_paint_time = segments[0].real_end_sand_time
            days_cost = 0.0
            now_day = 0
            for work in segments:
                cost_days, day_cost = divmod(work.paint, paint_efficiency)
                day_cost = round(day_cost / paint_efficiency, 2)
                days_cost += day_cost
                if days_cost >=1:
                    cost_days += 1
                    days_cost -= 1
                # work.real_end_sand_time = datetime.strptime(work.real_end_sand_time, '%Y/%m/%d').date()
                work.real_end_paint_time = timedelta(days=cost_days + now_day ) + start_paint_time
                now_day += cost_days
                work_info = [work.GC, work.ZD, work.SG, work.sand_area, work.paint_area,
                             work.ZYQ, team_name[workshop_idx], 2 * workshop_idx + group_idx,
                             work.real_end_sand_time,
                             work.CX, work.length, work.weight, work.height, work.n_paint,
                             work.pre_end_paint_time, work.real_end_paint_time]

                data_list.append(work_info)

    # 转换为 DataFrame
    df = pd.DataFrame(data_list)

    # 计算写入起始位置（如 "A1" 对应行=0, 列=0）
    start_row = int(insert_cell[1:]) - 1
    start_col = ord(insert_cell[0].upper()) - ord("A")

    # 写入 Excel
    with pd.ExcelWriter(new_file_path, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            index=False,
            header=False,
            startrow=start_row,
            startcol=start_col
        )
    data_center(new_file_path, insert_cell)